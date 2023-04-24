"""This file preprocesses data for a model which will predict whether a rodent is dependent (room air) or non-dependent (vapor) for each sex."""

import os

import numpy as np
import pandas
import pandas as pd
import scipy.signal
from scipy import signal
from itertools import combinations
from pypl2 import pl2_ad, pl2_info, pl2_comments
from statistics import fmean
import openpyxl
from API_Controller import *
from Load_Data import *


class Config:
    """This class holds info for the configuration of our data cleaning"""

    def __init__(self):
        self.filterRange = [57, 63]
        self.dwnSample = 5
        self.artifactThreshold = 1.5
        self.onset = 0.0125  # 25 values prior
        self.offset = 0.5  # 1000 values after
        self.sex = 'A'  # set to 'F' to process data for female models or 'M' for male models, or 'A' for all sexes
        self.excel_sheet = r'C:\Users\charl\Desktop\Data\RatData.xlsx'
        self.batches = 1  # set this value to 0 if you do not want batches, 1 if you do want batches

class AccessData:
    """Class used to access and process data from a directory of data files into power and coherence data that can be used
    in the program. The method getDataForLasso() can be called to populate data into a pandas dataframe located in this class.
    The class object will hold this dataframe."""

    def __init__(self, sDir, cfg):
        self.sDir = sDir
        self.cfg = cfg
        self.fType = '.pl2'  # File type we are looking for
        self.dFrameDict = {}
        self.dataframe = pandas.DataFrame()

        # Lists of channel names to be used in the below loops
        self.power_channel_names = ['Channel 1 Power', 'Channel 2 Power', 'Channel 3 Power', 'Channel 4 Power',
                                    'Channel 5 Power', 'Channel 6 Power']
        self.coherence_channel_names = ['Coherence 1 & 2', 'Coherence 1 & 3', 'Coherence 1 & 4', 'Coherence 1 & 5',
                                        'Coherence 1 & 6', 'Coherence 2 & 3', 'Coherence 2 & 4', 'Coherence 2 & 5',
                                        'Coherence 2 & 6', 'Coherence 3 & 4', 'Coherence 3 & 5', 'Coherence 3 & 6',
                                        'Coherence 4 & 5', 'Coherence 4 & 6', 'Coherence 5 & 6']
        self.band_names = [' Delta', ' Theta', ' Alpha', ' Beta', ' Low Gamma', ' High Gamma']

        if sDir != '':
            self.pl2_files = self.__getFileNames()

            self.preProcessData()

            self.header = []
            if self.cfg.batches == 1:
                # Equalize the length of each row in the dictionary
                self.__equalizeRows()

                # append the target feature to the end of each row
                self.__appendTargetFeature()

                # set the header and rows for the dataframe
                self.__buildHeader5SecondBatches()
            else:
                # set the header and rows for the dataframe
                self.__buildHeader()

            # Turn Dictionary dFrameDict into Dataframe
            self.dataframe = pd.DataFrame.from_dict(self.dFrameDict, orient='index', columns=self.header)
            print(self.dataframe.to_string())
            # print(self.dataframe.to_string())
            self.saveDataframe()

    def __appendTargetFeature(self):
        """Append Condition variable to end of array. 0 for Room Air, 1 for Vapor"""

        # open the Excel sheet with additional info
        wb = openpyxl.load_workbook(self.cfg.excel_sheet)
        ws = wb.active
        # iterate through all files and populate the pandas dataframe with power and coherence values
        for row in ws.iter_rows(values_only=True):
            if row[0]+'.pl2' in self.dFrameDict.keys():
                if row[3] == 'Room Air':
                    self.dFrameDict[row[0]+'.pl2'].append(0)
                if row[3] == 'Vapor':
                    self.dFrameDict[row[0]+'.pl2'].append(1)
    def __equalizeRows(self):
        """Since data files are of different lengths, we will find the shortest array and dock the rest to that length to avoid unequal column lengths"""
        # find the shortest array
        minimum = min(map(len, self.dFrameDict.values()))
        print("\n\n {} \n\n".format(minimum))

        # truncate all values in the dictionary to match the min length
        dictCopy = self.dFrameDict.copy()
        for key in dictCopy.keys():
            self.dFrameDict[key] = list(dictCopy[key])[0:minimum]

    def preProcessData(self):
        """Main method of this class. When called, it will populate the pandas dataframe stored in self. with all power
        and coherence values from the data files in the directory pointed at by this class."""
        print("Getting data for Lasso")
        os.chdir(self.sDir)  # change the current working directory to where our data is stored.

        # open the Excel sheet with additional info
        wb = openpyxl.load_workbook(self.cfg.excel_sheet)
        ws = wb.active

        counter = 0
        num_files = 0

        excel_files = []
        # count files to be read and check for duplicate names
        if self.cfg.sex == 'F' or self.cfg.sex == 'M':
            for row in ws.iter_rows(values_only=True):
                if row[0] not in excel_files:
                    excel_files.append(row[0])
                else:
                    print("Duplicate file name {} detected in excel spreadsheet".format(row[0]))
                    exit(1)
                if row[2] == self.cfg.sex and (row[0] + '.pl2') in self.pl2_files:
                    num_files += 1
        else:
            num_files = ws.max_row

        # iterate through all files and populate the pandas dataframe with power and coherence values
        for row in ws.iter_rows(values_only=True):
            if row[2] == self.cfg.sex or self.cfg.sex == 'A':
                for pl2_filename in self.pl2_files:
                    if pl2_filename == (row[0] + '.pl2'):
                        counter = counter + 1
                        print("\n   Working on file {}/{}".format(counter, num_files))
                        print("Processing file {} for Rat # {}".format(pl2_filename, row[1]))
                        if self.cfg.batches == 1:
                            self.__pl2ToDictionaryRow5SecondBatches(pl2_filename, row)
                        else:
                            self.__pl2ToDictionaryRow(pl2_filename, row)

    def saveDataframe(self):
        """Method that allows us to save our dataframe to an Excel spreadsheet. This way, we don't have to process the data every time we want to use it."""
        with pd.ExcelWriter('output.xlsx') as writer:
            self.dataframe.to_excel(writer)

    def __buildHeader5SecondBatches(self):
        """This function creates the rows and columns for our dataframe based on the amount of data in the 5-second batches."""
        header = []
        vals = list(self.dFrameDict.values())
        for i in range(len(vals[0])-1):
            header.append(i)

        header.append('Condition')
        print(header)
        self.header = header

    def __buildHeader(self):
        """This function creates the rows and columns for our dataframe. Since the header is 216 columns, it's much easier
        to create it programmatically"""

        # order is all powers + bandNames, then all coherences + band names
        header = []

        # power channels + bands
        for i in range(0, 6):
            for j in range(0, 6):
                header.append(self.power_channel_names[i] + self.band_names[j])

        # coherence channels + bands
        for i in range(0, 15):
            for j in range(0, 6):
                header.append(self.coherence_channel_names[i] + self.band_names[j])

        header.append('Condition')

        self.header = header

    def __splitSignal(self, f, signal):
        """splits an array of signal (power or coherence) values into an array of 6 arrays based on the 6 signal labels"""
        delta = []  # range 1-4
        theta = []  # range 5-10
        alpha = []  # range 11-14
        beta = []  # range 15-30
        low_gamma = []  # range 45-65
        high_gamma = []  # range 70-90

        for i in range(len(f)):
            if 1 <= f[i] <= 4:
                delta.append(signal[i])
            if 5 <= f[i] <= 10:
                theta.append(signal[i])
            if 11 <= f[i] <= 14:
                alpha.append(signal[i])
            if 15 <= f[i] <= 30:
                beta.append(signal[i])
            if 45 <= f[i] <= 65:
                low_gamma.append(signal[i])
            if 70 <= f[i] <= 90:
                high_gamma.append(signal[i])

        if len(delta) == 0:
            delta.append(0)
        if len(theta) == 0:
            theta.append(0)
        if len(alpha) == 0:
            alpha.append(0)
        if len(beta) == 0:
            beta.append(0)
        if len(low_gamma) == 0:
            low_gamma.append(0)
        if len(high_gamma) == 0:
            high_gamma.append(0)

        l = []
        l.append(fmean(delta))
        l.append(fmean(theta))
        l.append(fmean(alpha))
        l.append(fmean(beta))
        l.append(fmean(low_gamma))
        l.append(fmean(high_gamma))

        return l

    def __pl2ToDictionaryRow5SecondBatches(self, filename, row):
        """This method converts a pl2 file to a dictionary row of power and coherence data.
        Whereas the method __pl2ToDictionaryRow() converts the entire file's data into power and coherence,
        this method converts data in 5-second batches.
        The length of the row is returned so that we can build an appropriately sized header."""

        # count the number of active channels and store their numbers
        file_resource = pl2_info(filename)
        channel_count = 0
        channel_number = 0
        channel_numbers = []

        for channel in file_resource.ad:
            if channel.n > 0:  # if a channel has count > 0 then it has data
                channel_count += 1
                channel_numbers.append(channel_number)
            channel_number += 1  # update the current channel's number

        print("Detected {} active channels".format(channel_count))

        # convert data in channels from volts to raw a/d
        ad_array = []
        for channel in channel_numbers:
            print("Converting volts to a/d in channel {}".format(channel + 1))
            ad_info = pl2_ad(filename, channel)
            ad = self.voltsToRawAD(ad_info.ad)
            ad_array.append(ad)

        # perform data cleaning
        cleaned_ad_array = []
        iterator = 1
        for ad in ad_array:
            print("Cleaning Channel {}".format(iterator))

            # 1. 60 Hertz Filter
            print("Applying 60hz Filter")
            ad = self.__60HertzFilter(ad, ad_info.adfrequency)

            # 2. Down-sampling
            print("Applying Down-Sampling")
            ad = self.__downSampling(ad, self.cfg.dwnSample, ad_info.adfrequency)

            # 3. Threshold filter
            # print("Applying Threshold Filter")
            # ad = self.__noiseArtifactsFilter(ad, self.cfg.artifactThreshold, self.cfg.onset, self.cfg.offset, ad_info.adfrequency)

            cleaned_ad_array.append(ad)
            iterator += 1

        del ad_array  # release memory of uncleaned data

        # split data into 15-second batches
        sublist_ad_array = []
        num_sublists = int(ad_info.n / (ad_info.adfrequency * 60))  # There should be this many sublists
        for ad in cleaned_ad_array:
            sublist = np.array_split(ad, num_sublists)
            sublist_ad_array.append(sublist)

        # calculate power values for data
        power_array = []
        iterator = 0
        for channel in sublist_ad_array:
            print("Processing Power for Channel {}".format(self.power_channel_names[iterator]))
            iterator += 1
            channel_sublist = []
            for batch in channel:
                f, Pxx = self.__calculateChannelPower(batch, ad_info.adfrequency)
                channel_sublist.append((f, Pxx))
            power_array.append(channel_sublist)

        # calculate coherence values for data
        iterator = 0
        combo_list = list(combinations(channel_numbers, 2))  # get combinations of each 0-indexed channel number
        coherence_array = []
        for combo in combo_list:
            print("Processing Coherence for Channel Pair {}".format(self.coherence_channel_names[iterator]))
            iterator += 1
            ad1 = sublist_ad_array[combo[0]]
            ad2 = sublist_ad_array[combo[1]]
            channel_sublist = []
            for batch1, batch2 in zip(ad1, ad2):
                f, Cxy = self.__calculateChannelCoherence(batch1, batch2, ad_info.adfrequency)
                channel_sublist.append((f, Cxy))
            coherence_array.append(channel_sublist)

        # Split power signals into power bands and combine into one list
        split_signal_array = []
        for channel in power_array:
            channel_sublist = []
            for tup in channel:
                channel_sublist.append(self.__splitSignal(tup[0], tup[1]))
            split_signal_array.append(channel_sublist)

        # Split coherence signals into power bands and combine into one list
        for channel in coherence_array:
            channel_sublist = []
            for tup in channel:
                channel_sublist.append(self.__splitSignal(tup[0], tup[1]))
            split_signal_array.append(channel_sublist)

        # split_signal_array is in the order 5-second batches for channels 1-6 power, then 5-second batches for channel 1-6 coherences
        # convert split_signal_array into one list without any sublists
        combined_split_signal_array = []
        for l1 in split_signal_array:
            for l2 in l1:
                for l3 in l2:
                    combined_split_signal_array.append(l3)

        # Add info to dictionary
        self.dFrameDict[filename] = combined_split_signal_array

    def __pl2ToDictionaryRow(self, filename, row):
        """This method converts a pl2 file to a dictionary row of power and coherence data"""

        # count the number of active channels and store their numbers
        file_resource = pl2_info(filename)
        channel_count = 0
        channel_number = 0
        channel_numbers = []

        for channel in file_resource.ad:
            if channel.n > 0:  # if a channel has count > 0 then it has data
                channel_count += 1
                channel_numbers.append(channel_number)
            channel_number += 1  # update the current channel's number

        print("Detected {} active channels".format(channel_count))

        # convert data in channels from volts to raw a/d
        ad_array = []
        for channel in channel_numbers:
            print("Converting volts to a/d in channel {}".format(channel + 1))
            ad_info = pl2_ad(filename, channel)
            ad = self.voltsToRawAD(ad_info.ad)
            ad_array.append(ad)

        # perform data cleaning
        cleaned_ad_array = []
        iterator = 1
        for ad in ad_array:
            print("Cleaning Channel {}".format(iterator))

            # 1. 60 Hertz Filter
            print("Applying 60hz Filter")
            ad = self.__60HertzFilter(ad, ad_info.adfrequency)

            # 2. Down-sampling
            print("Applying Down-Sampling")
            ad = self.__downSampling(ad, self.cfg.dwnSample, ad_info.adfrequency)

            # 3. Threshold filter
            # print("Applying Threshold Filter")
            # ad = self.__noiseArtifactsFilter(ad, self.cfg.artifactThreshold, self.cfg.onset, self.cfg.offset, ad_info.adfrequency)

            cleaned_ad_array.append(ad)
            iterator += 1

        del ad_array  # release memory of uncleaned data

        # calculate power values for data
        power_array = []
        iterator = 0
        for ad in cleaned_ad_array:
            print("Processing Power for Channel {}".format(self.power_channel_names[iterator]))
            f, Pxx = self.__calculateChannelPower(ad, ad_info.adfrequency)
            power_array.append((f, Pxx))
            iterator += 1

        # calculate coherence values for data
        iterator = 0
        combo_list = list(combinations(channel_numbers, 2))  # get combinations of each 0-indexed channel number
        coherence_array = []
        for combo in combo_list:
            print("Processing Coherence for Channel Pair {}".format(self.coherence_channel_names[iterator]))
            ad1 = cleaned_ad_array[combo[0]]
            ad2 = cleaned_ad_array[combo[1]]
            f, Cxy = self.__calculateChannelCoherence(ad1, ad2, ad_info.adfrequency)
            coherence_array.append((f, Cxy))
            iterator += 1

        # Split power signals into power bands and combine into one list
        split_signal_array = []
        for tup in power_array:
            split_signal_array.extend(self.__splitSignal(tup[0], tup[1]))

        # Split coherence signals into power bands and combine into one list
        for tup in coherence_array:
            split_signal_array.extend(self.__splitSignal(tup[0], tup[1]))

        # Append Condition variable to end of array. 0 for Room Air, 1 for Vapor
        if row[3] == 'Room Air':
            split_signal_array.append(0)
        if row[3] == 'Vapor':
            split_signal_array.append(1)

        # Add info to dictionary
        self.dFrameDict[filename] = split_signal_array

    def __getFileNames(self):
        """Method for getting a list of file names from the directory pointed at by this class"""
        print("Getting file names")
        files = []

        for f in os.listdir(self.sDir):
            if f.endswith(self.fType):
                files.append(f)
        return files

    def __calculateChannelPower(self, ad, frequency):
        """Method for calculating the power values of a channel's LFP data. This function is trivial, but useful
        for making the code more readable. """
        f, Pxx = signal.welch(ad, fs=frequency)
        return f, Pxx

    def __calculateChannelCoherence(self, ad1, ad2, frequency):
        """Method for calculating the coherence values of a channel's LFP data. This function is trivial,
        but useful for making the code more readable. """
        f, Cxy = signal.coherence(ad1, ad2, fs=frequency)
        return f, Cxy

    def voltsToRawAD(self, arr):
        arr = list(map(lambda x: x / (1.9488 * pow(10, -7)), arr))
        arr = list(map(lambda x: round(x), arr))
        return arr

    def __60HertzFilter(self, sig, freq):
        """Cleans the data for frequencies of 60Hz using a second order Chebyshev type 1 notch filter. This is because the machine used to collect data naturally produces
        this signal, so it cannot be used. x is array to be cleaned, freq is sampling frequency."""

        b, a = scipy.signal.iirnotch(60, 30, freq)
        sig = scipy.signal.filtfilt(b, a, sig)
        return sig

    def __noiseArtifactsFilter(self, sig, artifactThreshold, onset, offset, f):
        """Cleans the data for noise artifacts created by interference by sources like the rat bashing its head against the enclosure wall.
        Filter is performed by scanning the signal array for values greater than the threshold, then removing all values a fraction of a second before and after that value."""

        ranges_to_remove = []  # make list of signals to remove after searching
        onset = round(onset * f)
        offset = round(offset * f)

        # iterate through array
        end_of_array = len(sig) - 1
        for i in range(len(sig)):
            if sig[i] >= artifactThreshold:  # if a value is greater than the threshold
                ranges_to_remove.extend(range(i - onset, i + offset + 1))

        ranges_to_remove = list(filter(lambda x: x >= 0, ranges_to_remove))
        ranges_to_remove = list(filter(lambda x: x <= end_of_array, ranges_to_remove))

        ranges_to_remove = list(set(ranges_to_remove))  # remove potential duplicates

        for r in sorted(ranges_to_remove, reverse=True):
            del sig[r]

        clean_sig = sig

        return clean_sig

    def __downSampling(self, sig, dsf, adfreq):
        """Downsamples the data for faster processing. sig is the signal to be downsampled. dsf needs to be a divisor of adfreq, which is the sampling frequency."""

        if adfreq % dsf != 0:
            raise Exception("Downsampling Frequency is not a divisor of Sampling Frequency")

        sig = scipy.signal.decimate(sig, dsf)

        return sig

if __name__ == "__main__":
    cfg = Config()
    configurator = API_Controller()
    # configurator.create_binary_config_file()
    configurator.update_binary_config("binary_config.ini", cfg)
    accessObj = AccessData(r'C:\Users\charl\Documents\SampleData', cfg)

    # loader = LoadData(r'D:\CS_421\Binary_Predictor_Data\output.xlsx')
    # loader.printDataFrame()

